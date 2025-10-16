#!/usr/bin/env python3
"""
Gerador de Planilha Excel Interativa das Carteiras Magnus.
Permite ao usuário inserir o valor total e calcula automaticamente a alocação.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def criar_excel_carteiras():
    """Cria planilha Excel interativa das carteiras."""
    
    filename = "Carteiras_Magnus_Outubro_2025.xlsx"
    wb = openpyxl.Workbook()
    
    # Remover sheet padrão
    wb.remove(wb.active)
    
    # Preços de referência (em reais)
    precos = {
        'IVVB11': 310.00,
        'LFTB11': 110.00,
        'BBAS3': 28.50,
        'BRSR6': 12.80,
        'BMGB4': 6.45,
        'PETR4': 38.20,
        'PRIO3': 45.60,
        'USIM5': 7.85,
        'GOAU4': 11.20,
        'BRAP4': 22.30,
        'LOGG3': 18.90,
        'DEXP3': 15.40,
        'SMTO3': 32.70,
        'TASA4': 9.15,
        'EUCA4': 4.25,
        'ALLD3': 8.90,
        'ROMI3': 24.50,
        'ITUB4': 32.50,
        'VALE3': 65.20,
        'WEGE3': 42.80
    }
    
    # ==================== CARTEIRA AGRESSIVA ====================
    ws_agressiva = wb.create_sheet("AGRESSIVA")
    
    # Título
    ws_agressiva.merge_cells('A1:G1')
    ws_agressiva['A1'] = '📊 CARTEIRA AGRESSIVA - OUTUBRO/2025'
    ws_agressiva['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_agressiva['A1'].fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
    ws_agressiva['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_agressiva.row_dimensions[1].height = 30
    
    # Informações
    ws_agressiva['A3'] = 'Perfil: Alta exposição a ações (46.67%)'
    ws_agressiva['A4'] = 'Risco: Alto | Retorno Esperado: 15-25% ao ano'
    ws_agressiva['A3'].font = Font(bold=True)
    ws_agressiva['A4'].font = Font(italic=True)
    
    # Campo de entrada
    ws_agressiva['A6'] = '💰 VALOR TOTAL A INVESTIR (R$):'
    ws_agressiva['A6'].font = Font(size=12, bold=True, color='FFFFFF')
    ws_agressiva['A6'].fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    ws_agressiva['B6'] = 20000  # Valor padrão
    ws_agressiva['B6'].font = Font(size=12, bold=True)
    ws_agressiva['B6'].fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    ws_agressiva['B6'].number_format = 'R$ #,##0.00'
    
    # Cabeçalho da tabela
    headers = ['Ticker', 'Nome', 'Setor', 'Alocação %', 'Valor (R$)', 'Preço (R$)', 'Qtd Aprox']
    for col, header in enumerate(headers, 1):
        cell = ws_agressiva.cell(8, col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='424242', end_color='424242', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados AGRESSIVA
    dados_agressiva = [
        ('IVVB11', 'iShares S&P 500', 'Internacional', 25.00),
        ('LFTB11', 'Tesouro Selic', 'Renda Fixa', 25.00),
        ('BBAS3', 'Banco do Brasil', 'Bancos', 3.33),
        ('BRSR6', 'Banrisul', 'Bancos', 3.33),
        ('BMGB4', 'Banco BMG', 'Bancos', 3.33),
        ('PETR4', 'Petrobras', 'Petróleo', 3.33),
        ('PRIO3', 'PetroRio', 'Petróleo', 3.33),
        ('USIM5', 'Usiminas', 'Siderurgia', 3.33),
        ('GOAU4', 'Gerdau Met', 'Siderurgia', 3.33),
        ('BRAP4', 'Bradespar', 'Mineração', 3.33),
        ('LOGG3', 'Log Commercial', 'Logística', 3.33),
        ('DEXP3', 'Dexxos', 'Logística', 3.33),
        ('SMTO3', 'São Martinho', 'Agronegócio', 3.33),
        ('TASA4', 'Taurus', 'Defesa', 3.33),
        ('EUCA4', 'Eucatex', 'Madeira', 3.33),
        ('ALLD3', 'Allied', 'Educação', 3.33),
        ('ROMI3', 'Romi', 'Máquinas', 3.33),
    ]
    
    row = 9
    for ticker, nome, setor, alocacao in dados_agressiva:
        ws_agressiva[f'A{row}'] = ticker
        ws_agressiva[f'B{row}'] = nome
        ws_agressiva[f'C{row}'] = setor
        ws_agressiva[f'D{row}'] = alocacao / 100
        ws_agressiva[f'D{row}'].number_format = '0.00%'
        
        # Fórmula para calcular valor
        ws_agressiva[f'E{row}'] = f'=$B$6*D{row}'
        ws_agressiva[f'E{row}'].number_format = 'R$ #,##0.00'
        
        # Preço
        ws_agressiva[f'F{row}'] = precos.get(ticker, 0)
        ws_agressiva[f'F{row}'].number_format = 'R$ #,##0.00'
        
        # Quantidade
        ws_agressiva[f'G{row}'] = f'=INT(E{row}/F{row})'
        ws_agressiva[f'G{row}'].number_format = '0'
        
        row += 1
    
    # Total
    ws_agressiva[f'D{row}'] = f'=SUM(D9:D{row-1})'
    ws_agressiva[f'D{row}'].number_format = '0.00%'
    ws_agressiva[f'D{row}'].font = Font(bold=True)
    ws_agressiva[f'E{row}'] = f'=SUM(E9:E{row-1})'
    ws_agressiva[f'E{row}'].number_format = 'R$ #,##0.00'
    ws_agressiva[f'E{row}'].font = Font(bold=True)
    ws_agressiva[f'A{row}'] = 'TOTAL'
    ws_agressiva[f'A{row}'].font = Font(bold=True)
    
    # Ajustar larguras
    ws_agressiva.column_dimensions['A'].width = 12
    ws_agressiva.column_dimensions['B'].width = 20
    ws_agressiva.column_dimensions['C'].width = 15
    ws_agressiva.column_dimensions['D'].width = 12
    ws_agressiva.column_dimensions['E'].width = 15
    ws_agressiva.column_dimensions['F'].width = 12
    ws_agressiva.column_dimensions['G'].width = 12
    
    # ==================== CARTEIRA MODERADA ====================
    ws_moderada = wb.create_sheet("MODERADA")
    
    # Título
    ws_moderada.merge_cells('A1:G1')
    ws_moderada['A1'] = '📊 CARTEIRA MODERADA - OUTUBRO/2025'
    ws_moderada['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_moderada['A1'].fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
    ws_moderada['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_moderada.row_dimensions[1].height = 30
    
    # Informações
    ws_moderada['A3'] = 'Perfil: Balanceada (25% ações)'
    ws_moderada['A4'] = 'Risco: Médio | Retorno Esperado: 10-15% ao ano'
    ws_moderada['A3'].font = Font(bold=True)
    ws_moderada['A4'].font = Font(italic=True)
    
    # Campo de entrada
    ws_moderada['A6'] = '💰 VALOR TOTAL A INVESTIR (R$):'
    ws_moderada['A6'].font = Font(size=12, bold=True, color='FFFFFF')
    ws_moderada['A6'].fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    ws_moderada['B6'] = 20000
    ws_moderada['B6'].font = Font(size=12, bold=True)
    ws_moderada['B6'].fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    ws_moderada['B6'].number_format = 'R$ #,##0.00'
    
    # Cabeçalho
    for col, header in enumerate(headers, 1):
        cell = ws_moderada.cell(8, col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='424242', end_color='424242', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados MODERADA (mesmos ativos, alocações diferentes)
    dados_moderada = [
        ('LFTB11', 'Tesouro Selic', 'Renda Fixa', 50.00),
        ('IVVB11', 'iShares S&P 500', 'Internacional', 25.00),
        ('BBAS3', 'Banco do Brasil', 'Bancos', 1.67),
        ('BRSR6', 'Banrisul', 'Bancos', 1.67),
        ('BMGB4', 'Banco BMG', 'Bancos', 1.67),
        ('PETR4', 'Petrobras', 'Petróleo', 1.67),
        ('PRIO3', 'PetroRio', 'Petróleo', 1.67),
        ('USIM5', 'Usiminas', 'Siderurgia', 1.67),
        ('GOAU4', 'Gerdau Met', 'Siderurgia', 1.67),
        ('BRAP4', 'Bradespar', 'Mineração', 1.67),
        ('LOGG3', 'Log Commercial', 'Logística', 1.67),
        ('DEXP3', 'Dexxos', 'Logística', 1.67),
        ('SMTO3', 'São Martinho', 'Agronegócio', 1.67),
        ('TASA4', 'Taurus', 'Defesa', 1.67),
        ('EUCA4', 'Eucatex', 'Madeira', 1.67),
        ('ALLD3', 'Allied', 'Educação', 1.67),
        ('ROMI3', 'Romi', 'Máquinas', 1.67),
    ]
    
    row = 9
    for ticker, nome, setor, alocacao in dados_moderada:
        ws_moderada[f'A{row}'] = ticker
        ws_moderada[f'B{row}'] = nome
        ws_moderada[f'C{row}'] = setor
        ws_moderada[f'D{row}'] = alocacao / 100
        ws_moderada[f'D{row}'].number_format = '0.00%'
        ws_moderada[f'E{row}'] = f'=$B$6*D{row}'
        ws_moderada[f'E{row}'].number_format = 'R$ #,##0.00'
        ws_moderada[f'F{row}'] = precos.get(ticker, 0)
        ws_moderada[f'F{row}'].number_format = 'R$ #,##0.00'
        ws_moderada[f'G{row}'] = f'=INT(E{row}/F{row})'
        ws_moderada[f'G{row}'].number_format = '0'
        row += 1
    
    # Total
    ws_moderada[f'D{row}'] = f'=SUM(D9:D{row-1})'
    ws_moderada[f'D{row}'].number_format = '0.00%'
    ws_moderada[f'D{row}'].font = Font(bold=True)
    ws_moderada[f'E{row}'] = f'=SUM(E9:E{row-1})'
    ws_moderada[f'E{row}'].number_format = 'R$ #,##0.00'
    ws_moderada[f'E{row}'].font = Font(bold=True)
    ws_moderada[f'A{row}'] = 'TOTAL'
    ws_moderada[f'A{row}'].font = Font(bold=True)
    
    # Ajustar larguras
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_moderada.column_dimensions[col].width = ws_agressiva.column_dimensions[col].width
    
    # ==================== CARTEIRA CONSERVADORA ====================
    ws_conservadora = wb.create_sheet("CONSERVADORA")
    
    # Título
    ws_conservadora.merge_cells('A1:G1')
    ws_conservadora['A1'] = '🛡️ CARTEIRA CONSERVADORA - OUTUBRO/2025'
    ws_conservadora['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_conservadora['A1'].fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
    ws_conservadora['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_conservadora.row_dimensions[1].height = 30
    
    # Informações
    ws_conservadora['A3'] = 'Perfil: Baixa exposição a ações (10%)'
    ws_conservadora['A4'] = 'Risco: Baixo | Retorno Esperado: 8-12% ao ano'
    ws_conservadora['A3'].font = Font(bold=True)
    ws_conservadora['A4'].font = Font(italic=True)
    
    # Campo de entrada
    ws_conservadora['A6'] = '💰 VALOR TOTAL A INVESTIR (R$):'
    ws_conservadora['A6'].font = Font(size=12, bold=True, color='FFFFFF')
    ws_conservadora['A6'].fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    ws_conservadora['B6'] = 15000
    ws_conservadora['B6'].font = Font(size=12, bold=True)
    ws_conservadora['B6'].fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    ws_conservadora['B6'].number_format = 'R$ #,##0.00'
    
    # Cabeçalho
    for col, header in enumerate(headers, 1):
        cell = ws_conservadora.cell(8, col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='424242', end_color='424242', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados CONSERVADORA
    dados_conservadora = [
        ('LFTB11', 'Tesouro Selic', 'Renda Fixa', 70.00),
        ('IVVB11', 'iShares S&P 500', 'Internacional', 20.00),
        ('BBAS3', 'Banco do Brasil', 'Bancos', 2.00),
        ('ITUB4', 'Itaú Unibanco', 'Bancos', 2.00),
        ('PETR4', 'Petrobras', 'Petróleo', 2.00),
        ('VALE3', 'Vale', 'Mineração', 2.00),
        ('WEGE3', 'WEG', 'Equipamentos', 2.00),
    ]
    
    row = 9
    for ticker, nome, setor, alocacao in dados_conservadora:
        ws_conservadora[f'A{row}'] = ticker
        ws_conservadora[f'B{row}'] = nome
        ws_conservadora[f'C{row}'] = setor
        ws_conservadora[f'D{row}'] = alocacao / 100
        ws_conservadora[f'D{row}'].number_format = '0.00%'
        ws_conservadora[f'E{row}'] = f'=$B$6*D{row}'
        ws_conservadora[f'E{row}'].number_format = 'R$ #,##0.00'
        ws_conservadora[f'F{row}'] = precos.get(ticker, 0)
        ws_conservadora[f'F{row}'].number_format = 'R$ #,##0.00'
        ws_conservadora[f'G{row}'] = f'=INT(E{row}/F{row})'
        ws_conservadora[f'G{row}'].number_format = '0'
        row += 1
    
    # Total
    ws_conservadora[f'D{row}'] = f'=SUM(D9:D{row-1})'
    ws_conservadora[f'D{row}'].number_format = '0.00%'
    ws_conservadora[f'D{row}'].font = Font(bold=True)
    ws_conservadora[f'E{row}'] = f'=SUM(E9:E{row-1})'
    ws_conservadora[f'E{row}'].number_format = 'R$ #,##0.00'
    ws_conservadora[f'E{row}'].font = Font(bold=True)
    ws_conservadora[f'A{row}'] = 'TOTAL'
    ws_conservadora[f'A{row}'].font = Font(bold=True)
    
    # Ajustar larguras
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws_conservadora.column_dimensions[col].width = ws_agressiva.column_dimensions[col].width
    
    # Salvar
    wb.save(filename)
    print(f"✅ Excel gerado: {filename}")
    return filename


if __name__ == '__main__':
    print("=" * 80)
    print("GERANDO PLANILHA EXCEL INTERATIVA")
    print("=" * 80)
    filename = criar_excel_carteiras()
    print(f"\n✅ Excel criado com sucesso: {filename}")
    print("\n💡 Como usar:")
    print("1. Abra o arquivo no Excel ou Google Sheets")
    print("2. Escolha a aba da carteira desejada")
    print("3. Digite o valor total a investir na célula amarela (B6)")
    print("4. A planilha calculará automaticamente quanto alocar em cada ativo")
    print("=" * 80)

